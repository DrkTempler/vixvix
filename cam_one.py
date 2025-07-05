from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from pymediainfo import MediaInfo
from openpyxl import Workbook, load_workbook
import requests
import time
import logging
import os
import sys
import io

# 터미널에 로그출력
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 메인페이지
def setup_driver():
    options = Options()
    options.add_experimental_option("detach", True)
    options.add_argument("--headless") 
    # options.add_argument('--start-maximized')
    service = Service()
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(3)
    return driver
# 로그인
def login(ip, driver, username, password):
    driver.get(ip)
    action = ActionChains(driver)
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[1]/div[2]/div/form/p[1]/input').click()
    action.send_keys(username).key_down(Keys.TAB).send_keys(password).pause(1).key_down(Keys.ENTER).perform()
    action.reset_actions()
    time.sleep(5)
    logger.info('login success')
# 설정로그인 + 비디오 스트림 3, 4 h.264로 설정
def loginop(ip, driver, username, password):
    driver.get(ip)
    time.sleep(5)
    try:
        element = driver.find_element(By.CSS_SELECTOR, '#user')
        element.click()
        element.send_keys(username)
        time.sleep(3)
        element = driver.find_element(By.CSS_SELECTOR, '#pwd')
        element.click()
        element.send_keys(password)
        time.sleep(3)
        element = driver.find_element(By.CSS_SELECTOR, '#login_form > p:nth-child(8) > div > ins')
        element.click()
        element = driver.find_element(By.CSS_SELECTOR, '#login_submit')
        element.click()
    except Exception as e:
        print(f"Error: {e}")
    time.sleep(5)
    logger.info('option menu login complete')
    time.sleep(5)
    try:
        element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li:nth-child(2)')
        element.click()
        time.sleep(3)
        element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li.open > ul > li:nth-child(2) > a')
        element.click()
        time.sleep(3)
        select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_codec_list'))
        select.select_by_value("H264HP")
        time.sleep(3)
        select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_codec_list'))
        select.select_by_value("H264HP")
        time.sleep(3)
        element = driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]')
        element.click()
    except Exception as e:
        print(f"Error: {e}")
    logger.info('Stream 3, 4 Codec Change Complete')

# 탭 호출 > 라이브
def tab_l(driver):
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[1]/div/button').click()
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[2]/div[1]/div[1]/div[1]/a[1]').click()
    logger.info('Go to Live Mode')

# 탭 호출 > 플레이백
def tab_p(driver):
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[1]/div/button').click()
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[2]/div[1]/div[1]/div[1]/a[2]').click()
    logger.info('Go to Playback Mode')

# 탭 호출 > 설정
def tab_o(driver):
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[1]/div/button').click()
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[2]/div[1]/div[1]/div[1]/a[3]').click()
    logger.info('Go to Options')

# 탭 호출 > 로그아웃
def tab_bye(driver):
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[1]/div/button').click()
    driver.find_element(By.CSS_SELECTOR, '/html/body/div[2]/div[1]/div[1]/div[1]/a[4]').click()
    logger.info('Logout Complete')

# 설정 > 비디오 스트림 진입
def op_stream(driver):
    try:
        element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li:nth-child(2)')
        element.click()
        time.sleep(3)
        element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li.open > ul > li:nth-child(2) > a')
        element.click()
        time.sleep(3)
    except Exception as e:
        print(f"Error: {e}")
    logger.info('Options > Video Stream Now')

# 플레이백 > 비디오 스트림 진입
def op_stream_pb(driver):
    try:
        element = driver.find_element(By.XPATH, '//*[@id="menu_button"]')
        element.click()
        element = driver.find_element(By.XPATH, '//*[@id="sidebar-shortcuts-large"]/a[3]')
        element.click()
        time.sleep(3)
        element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li:nth-child(2)')
        element.click()
        time.sleep(3)
        element = driver.find_element(By.CSS_SELECTOR, '#sidebar > ul > li.open > ul > li:nth-child(2) > a')
        element.click()
        time.sleep(3)
    except Exception as e:
        print(f"Error: {e}")
    logger.info('Playback > Video Stream Now')


# 저장소 겹쳐쓰기 설정
def reckeep(driver):
    element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[4]/a')
    element.click()
    element = driver.find_element(By.XPATH, '//*[@id="tab-record"]/div[1]/label/div')
    element.click()
    element = driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div/div[2]/div[1]/h1/button')
    element.click()
    logger.info('Keep Save Mode On')

# 이벤트 녹화 사용설정-비디오 스트림1
def recon(driver):
    try:
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/a')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/ul/li[2]/a')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="tab-record"]/div[1]/label/div/ins')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div/div[2]/div[1]/h1/button')
        element.click()
        time.sleep(5)
    except Exception as e:
        print(f"Error:{e}")
    logger.info('Event Rec On with Video Stream1')

# 이벤트 녹화 사용설정-비디오 스트림2
def recon2(driver):
    try:
        element = driver.find_element(By.XPATH, '//*[@id="menu_button"]')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar-shortcuts-large"]/a[3]')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/a')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/ul/li[2]/a')
        element.click()
        time.sleep(5)
        select = Select(driver.find_element(By.CSS_SELECTOR,'#real_record_event_video'))
        select.select_by_value('2')
        element = driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div/div[2]/div[1]/h1/button')
        element.click()
        time.sleep(5)
    except Exception as e:
        print(f"Error:{e}")
    logger.info('Event Rec On with Video Stream2')

# 이벤트 녹화 사용설정-비디오 스트림3
def recon3(driver):
    try:
        # element = driver.find_element(By.XPATH, '//*[@id="menu_button"]')
        # element.click()
        # time.sleep(5)
        # element = driver.find_element(By.XPATH, '//*[@id="sidebar-shortcuts-large"]/a[3]')
        # element.click()
        # time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/a')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/ul/li[2]/a')
        element.click()
        time.sleep(5)
        select = Select(driver.find_element(By.CSS_SELECTOR,'#real_record_event_video'))
        select.select_by_value('3')
        element = driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div/div[2]/div[1]/h1/button')
        element.click()
        time.sleep(5)
    except Exception as e:
        print(f"Error:{e}")
    logger.info('Event Rec On with Video Stream3')

# 이벤트 녹화 사용설정-비디오 스트림4
def recon4(driver):
    try:
        # element = driver.find_element(By.XPATH, '//*[@id="menu_button"]')
        # element.click()
        # time.sleep(5)
        # element = driver.find_element(By.XPATH, '//*[@id="sidebar-shortcuts-large"]/a[3]')
        # element.click()
        # time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/a')
        element.click()
        time.sleep(5)
        element = driver.find_element(By.XPATH, '//*[@id="sidebar"]/ul/li[5]/ul/li[2]/a')
        element.click()
        time.sleep(5)
        select = Select(driver.find_element(By.CSS_SELECTOR,'#real_record_event_video'))
        select.select_by_value('4')
        element = driver.find_element(By.XPATH, '//*[@id="main-container"]/div[2]/div/div[2]/div[1]/h1/button')
        element.click()
        time.sleep(5)
    except Exception as e:
        print(f"Error:{e}")
    logger.info('Event Rec On with Video Stream4')

# 이벤트 녹화 설정(타이머)
def rectimer(driver):
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/a'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/ul/li[1]/a'))).click()
    time.sleep(5)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[1]/ul/li[9]/a'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tab-timer"]/div[1]/label/div'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[3]/select'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[3]/select/option[1]'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[4]/select'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[9]/div[4]/select/option[2]'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[1]/h1/button'))).click()
    logger.info('Timer On(1m)')
    time.sleep(5)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[1]/ul/li[5]/ul/li[3]/a'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div[2]/form/div/div[2]/div[1]/div/div[2]/input[1]'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div[2]/form/div[2]/div[2]/div/div[1]/select'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div[2]/form/div[2]/div[2]/div/div[1]/select/option[46]'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="modal_rule"]/div/div/div[2]/form/div[3]/div[2]/div[8]/div/label/div'))).click()
    time.sleep(3)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div[3]/button[1]'))).click()
    logger.info('Create Event Rule with Timer')

# 이벤트 영상 추출하기
def videosave(driver, videoname):
    driver.find_element(By.XPATH, '//*[@id="sidebar-shortcuts-large"]/a[2]').click()
    rows = driver.find_elements(By.XPATH, '//tr[@role="row" and @id]')
    nowid = max(int(row.get_attribute('id')) for row in rows)
    elements = driver.find_elements(By.XPATH, f'//tr[@role="row" and @id="{nowid}"]/td[4]')
    if elements:
        elements[0].click()
    else:
        print("WHAT THE FUCK")
    video_element = driver.find_element(By.TAG_NAME, 'video')
    video_url = video_element.get_attribute('src')
    response = requests.get(video_url, timeout=60)
    directory = "D:/testvideo"
    os.makedirs(directory, exist_ok=True)
    file_name = f"{videoname}.mp4"
    file_path = os.path.join(directory, file_name)
    file = open(file_path, 'wb') 
    file.write(response.content)
    logger.info("File saved as %s", file_name)


#엑셀 저장
def excelsave(file_path, sheet_name, cell, log_message):
 
    if not os.path.exists(file_path):
        if not os.path.exists(os.path.dirname(file_path)):
         os.makedirs(os.path.dirname(file_path))
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws[cell] = log_message
        wb.save(file_path)
        print("result file on %s", file_path)
    else:
        print(f"already have directory, just only file")
        workbook = load_workbook(file_path)  
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook.create_sheet(sheet_name)  
        ws[cell] = log_message
        workbook.save(file_path) 
        workbook.close() 


# 이벤트 영상 조건 확인하기
def info(videoname):
    file_path = os.path.join("D:\\", "testvideo", f"{videoname}.mp4")
    media_info = MediaInfo.parse(file_path)
    result = []
    for track in media_info.tracks:
        if track.track_type == 'Video':
            result.append(f"코덱 프로파일, {track.format_profile}")
            result.append(f"영상길이, {track.duration}")
            result.append(f"fps, {track.frame_rate}")
            result.append(f"width, {track.width}")
            result.append(f"height, {track.height}")
            result.append(f"비트레이트 모드, {track.bit_rate_mode}")
            result.append(f"비트레이트, {track.bit_rate}")
    return "\n".join(result)

#cama-1
def cama1(driver):
    videoname = 'cama1'
    # 압축 방식 H.264 High
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_codec_list'))
    select.select_by_value("H264HP")
    time.sleep(3)
    # 해상도 2160
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_resolution_list'))
    select.select_by_value('3840x2160')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 6000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bps_list'))
    select.select_by_value('6000')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-1 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E3", log_message = info(videoname))

#cama-2
def cama2(driver):
    videoname = 'cama2'
    # 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 1728
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_resolution_list'))
    select.select_by_value('3072x1728')
    time.sleep(3)
    # 프레임레이트 15
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_fps_list'))
    select.select_by_value('15')
    time.sleep(3)
    # GOP크기 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_gov_list'))
    select.select_by_value('10')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 12000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bps_list'))
    select.select_by_value('12000')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-2 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E4", log_message = info(videoname))

# cama-3 압축방식 Smart
def cama3(driver):
    videoname = 'cama3'
    # 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 1440
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_resolution_list'))
    select.select_by_value('2560x1440')
    time.sleep(3)
    # 프레임레이트 1
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_fps_list'))
    select.select_by_value('1')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bit_control_list'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 1000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bps_list'))
    select.select_by_value('1000')
    time.sleep(3)
    # 비트레이트 품질 10
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc1_bps_quality > select'))
    select.select_by_value('10')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-3 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E5", log_message = info(videoname))

# cama-4 
def cama4(driver):
    videoname = 'cama4'
    # 압축 방식 H.264 Smart
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_codec_list'))
    select.select_by_value("H264PL")
    time.sleep(3)
    # 해상도 1440
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_resolution_list'))
    select.select_by_value('2560x1440')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_gov_list'))
    select.select_by_value('10')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bit_control_list'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 5000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc1_bps_list'))
    select.select_by_value('5000')
    time.sleep(3)
    # 비트레이트 품질 1
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc1_bps_quality > select'))
    select.select_by_value('1')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-4 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E6", log_message = info(videoname))

# cama-5 ************************비디오스트림2 필수지정!!*******************************
def cama5(driver):
    videoname = 'cama5'
    # 압축 방식 H.264 High
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264HP")
    time.sleep(3)
    # 해상도 1080
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('1920x1080')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 12000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('12000')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-5 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E7", log_message = info(videoname))

# cama-6
def cama6(driver):
    videoname = 'cama6'
# 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 896
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('1600x896')
    time.sleep(3)
    # 프레임레이트 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('10')
    time.sleep(3)
    # GOP크기 1
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('1')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 7500
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('7500')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-6 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E8", log_message = info(videoname))

# cama-7
def cama7(driver):
    videoname = 'cama7'
    # 압축 방식 H.264 Smart
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264PL")
    time.sleep(3)
    # 해상도 720
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('1280x720')
    time.sleep(3)
    # 프레임레이트 1
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('1')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 100
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('100')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-7 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E9", log_message = info(videoname))

# cama-8
def cama8(driver):
    videoname = 'cama8'
    # 압축 방식 H.264 MAIN
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 720
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('768x432')
    time.sleep(3)
    # 프레임레이트 1
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('1')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 100
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('100')
    time.sleep(3)
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-8 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E10", log_message = info(videoname))

# cama-9
def cama9(driver):
    videoname = 'cama9'
    # 압축 방식 H.264 MAIN
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 392
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('704x392')
    time.sleep(3)
    # 프레임레이트 20
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('20')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 1000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('1000')
    time.sleep(3)
    # 비트레이트 품질 1
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc2_bps_quality > select'))
    select.select_by_value('1')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-9 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E11", log_message = info(videoname))

# cama-10
def cama10(driver):
    videoname = 'cama10'
    # 압축 방식 H.264 MAIN
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 360
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_resolution_list'))
    select.select_by_value('640x360')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bit_control_list'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 3000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc2_bps_list'))
    select.select_by_value('3000')
    time.sleep(3)
    # 비트레이트 품질 10
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc2_bps_quality > select'))
    select.select_by_value('10')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-10 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E12", log_message = info(videoname))

# cama-11 !!!!!비디오 스트림 3!!!!!
def cama11(driver):
    videoname = 'cama11'
    # 압축 방식 H.264 High
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_codec_list'))
    select.select_by_value("H264HP")
    time.sleep(3)
    # 해상도 432
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_resolution_list'))
    select.select_by_value('768x432')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bit_control'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 8000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bps_list'))
    select.select_by_value('8000')
    time.sleep(3)
    # 비트레이트 품질 10
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc3_bps_quality > select'))
    select.select_by_value('10')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-11 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E13", log_message = info(videoname))

# cama-12
def cama12(driver):
    videoname = 'cama12'
    # 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 392
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_resolution_list'))
    select.select_by_value('704x392')
    time.sleep(3)
    # 프레임레이트 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_fps_list'))
    select.select_by_value('10')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bit_control'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 8000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bps_list'))
    select.select_by_value('1000')
    time.sleep(3)
    # 비트레이트 품질 10
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc3_bps_quality > select'))
    select.select_by_value('1')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-12 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E14", log_message = info(videoname))

# cama-13
def cama13(driver):
    videoname = 'cama13'
    # 압축 방식 H.264 Smart
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_codec_list'))
    select.select_by_value("H264PL")
    time.sleep(3)
    # 해상도 360
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_resolution_list'))
    select.select_by_value('640x360')
    time.sleep(3)
    # 프레임레이트 15
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_fps_list'))
    select.select_by_value('15')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bit_control'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 800
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc3_bps_list'))
    select.select_by_value('800')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-13 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E15", log_message = info(videoname))

# cama-14  !!!!!!!비디오스트림4!!!!!!!!!!!
def cama14(driver):
    videoname = 'cama14'
    # 압축 방식 H.264 High
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_codec_list'))
    select.select_by_value("H264HP")
    time.sleep(3)
    # 해상도 720
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_resolution_list'))
    select.select_by_value('1280x720')
    time.sleep(3)
    # 프레임레이트 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_fps_list'))
    select.select_by_value('30')
    time.sleep(3)
    # GOP크기 60
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_gov_list'))
    select.select_by_value('60')
    time.sleep(3)
    # 비트레이트 제어 CBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bit_control'))
    select.select_by_value('cbr')
    time.sleep(3)
    # 비트레이트 12000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bps_list'))
    select.select_by_value('12000')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-14 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E16", log_message = info(videoname))

def cama15(driver):
    videoname = 'cama15'
    # 압축 방식 H.264 High
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 432
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_resolution_list'))
    select.select_by_value('768x432')
    time.sleep(3)
    # 프레임레이트 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_fps_list'))
    select.select_by_value('10')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bit_control'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 1000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bps_list'))
    select.select_by_value('1000')
    # 비트레이트 품질 1
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc4_bps_quality > select'))
    select.select_by_value('1')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-15 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E17", log_message = info(videoname))

def cama16(driver):
    videoname = 'cama16'
    # 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 392
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_resolution_list'))
    select.select_by_value('704x392')
    time.sleep(3)
    # 프레임레이트 1
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_fps_list'))
    select.select_by_value('1')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bit_control'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 5000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bps_list'))
    select.select_by_value('5000')
    # 비트레이트 품질 10
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc4_bps_quality > select'))
    select.select_by_value('10')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-16 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E18", log_message = info(videoname))

def cama17(driver):
    videoname = 'cama17'
    # 압축 방식 H.264 Main
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_codec_list'))
    select.select_by_value("H264MP")
    time.sleep(3)
    # 해상도 360
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_resolution_list'))
    select.select_by_value('640x360')
    time.sleep(3)
    # 프레임레이트 10
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_fps_list'))
    select.select_by_value('10')
    time.sleep(3)
    # GOP크기 30
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_gov_list'))
    select.select_by_value('30')
    time.sleep(3)
    # 비트레이트 제어 VBR
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bit_control'))
    select.select_by_value('vbr')
    time.sleep(3)
    # 비트레이트 1000
    select = Select(driver.find_element(By.CSS_SELECTOR,'#video_enc4_bps_list'))
    select.select_by_value('1000')
    # 비트레이트 품질 1
    select = Select(driver.find_element(By.CSS_SELECTOR, '#row_video_enc4_bps_quality > select'))
    select.select_by_value('1')
    # 저장
    driver.find_element(By.XPATH, '//*[@id="btnSaveSetup"]').click()
    logger.info('CAMa-17 TC Precondition Set Complete')
    time.sleep(150)
    videosave(driver, videoname)
    file_path = "D:\\Auto_test\\VIXcam_AutoTC.xlsx"
    excelsave(file_path, sheet_name="VIXcam_AutoTC", cell="E19", log_message = info(videoname))

# # 테스트 진행 전 무조건 저장소 초기화, 시스템 초기화 후 비밀번호 초기 설정까지 완료하세요
def main_rec(ip, username, password):
    driver = setup_driver()
    loginop(ip, driver, username, password)
    time.sleep(5)
    recon(driver)
    time.sleep(5)
    reckeep(driver)
    time.sleep(5)
    rectimer(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama1(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama2(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama3(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama4(driver)
    time.sleep(5)
    recon2(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama5(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)    
    cama6(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama7(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama8(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama9(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama10(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    recon3(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama11(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama12(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama13(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    recon4(driver)
    time.sleep(5)
    op_stream(driver)
    time.sleep(5)
    cama14(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama15(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama16(driver)
    time.sleep(5)
    op_stream_pb(driver)
    time.sleep(5)
    cama17(driver)
    time.sleep(5)

if __name__ == "__main__":
    main_rec()